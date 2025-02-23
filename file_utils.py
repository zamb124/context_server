# --- START OF FILE file_utils.py ---
import asyncio
import functools
import io
import json

import docx
import openpyxl
import pdfplumber


def run_sync(func):
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, functools.partial(func, *args, **kwargs))

    return wrapper


@run_sync
def extract_text_from_pdf_sync(file: io.BytesIO) -> str:
    """Извлекает текст из PDF-файла (синхронная версия)."""
    with pdfplumber.open(file) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text


async def extract_text_from_pdf(file: io.BytesIO) -> str:
    """Извлекает текст из PDF-файла."""
    return await asyncio.to_thread(extract_text_from_pdf_sync, file)


# Аналогично для других функций в file_utils.py:
@run_sync
def extract_text_from_docx_sync(file: io.BytesIO) -> str:
    """Извлекает текст из DOCX-файла (синхронная версия)."""
    doc = docx.Document(file)
    text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
    return text


async def extract_text_from_docx(file: io.BytesIO) -> str:
    """Извлекает текст из DOCX-файла."""
    return await asyncio.to_thread(extract_text_from_docx_sync, file)


@run_sync
def extract_text_from_xlsx_sync(file: io.BytesIO) -> str:
    """Извлекает текст из XLSX-файла (синхронная версия)."""
    workbook = openpyxl.load_workbook(file)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            row_values = [str(cell.value) for cell in row if
                          cell.value is not None]
            text += ", ".join(row_values) + "\n"
    return text


async def extract_text_from_xlsx(file: io.BytesIO) -> str:
    """Извлекает текст из XLSX-файла."""
    return await asyncio.to_thread(extract_text_from_xlsx_sync, file)


async def extract_text_from_txt(file: io.BytesIO) -> str:
    """Извлекает текст из TXT-файла."""
    data = await asyncio.to_thread(file.read)
    return data.decode("utf-8")


async def extract_text_from_json(file: io.BytesIO) -> str:
    """Извлекает текст из JSON-файла (предполагается, что это просто текстовый JSON)."""
    try:
        data = await asyncio.to_thread(json.load, file)
        return json.dumps(data, indent=4)  # Форматируем JSON для читаемости
    except json.JSONDecodeError:
        raise ValueError("Некорректный JSON формат.")
