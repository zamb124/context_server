# --- Дополнительные библиотеки ---
import asyncio
import io
import json
import logging
import os
import re
import signal
# --- Дополнительные библиотеки для работы с Telegram ---
import ssl
import sys
import traceback
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Optional, Any

import aiofiles
import certifi
import chromadb
import docx
import nltk
import openpyxl
import pdfplumber
import uvicorn
from chromadb.utils import embedding_functions
from fastapi import FastAPI, HTTPException, Depends, Header, status, UploadFile, File
from langchain.text_splitter import RecursiveCharacterTextSplitter
from pydantic import BaseModel

from config import config  # Убедитесь, что config.py содержит CHAT_TOKEN и TELEGRAM_BOT_TOKEN

# --- Конфигурация ---
CHROMA_DB_PATH = "chroma_db"  # Папка, где будет храниться база данных ChromaDB
EMBEDDING_MODEL_NAME = "sentence-transformers/paraphrase-multilingual-mpnet-base-v2"  # ИЗМЕНЕНО
TEMP_STORAGE_PATH = "temp_telegram_data"  # Папка для временного хранения данных Telegram
SAVE_INTERVAL_SECONDS = 600  # Интервал сохранения в секундах (10 минут)
OFFSET_FILE = "offset.txt"

# --- Токен авторизации ---
CHAT_TOKEN = config.CHAT_TOKEN
if not CHAT_TOKEN:
    raise ValueError("Не установлена переменная окружения CHAT_TOKEN. Пожалуйста, настройте ее в файле .env.")

TELEGRAM_BOT_TOKEN = config.TELEGRAM_BOT_TOKEN
if not TELEGRAM_BOT_TOKEN:
    raise ValueError("Не установлена переменная окружения TELEGRAM_BOT_TOKEN. Пожалуйста, настройте ее в файле .env.")

HUBSPOT_API_KEY = config.HUBSPOT_API_KEY
if not HUBSPOT_API_KEY:
    raise ValueError("Не установлена переменная окружения HUBSPOT_API_KEY. Пожалуйста, настройте ее в файле .env.")

# --- Инициализация логирования ---
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


# --- Модели данных ---
class DocumentBase(BaseModel):
    id: str  # ID документа
    text: str
    label: str  # Обязательный атрибут label
    metadata: Dict[str, str] = {}  # Добавляем метаданные


class Query(BaseModel):
    text: str
    labels: List[str]  # Обязательный атрибут labels
    n_results: int = 5
    where: Optional[Dict[str, Any]] = None  # Фильтры по метаданным


class ContextResponse(BaseModel):
    results: List[Dict]


class ForceSaveResponse(BaseModel):
    message: str


# --- Инициализация ChromaDB ---
sentence_transformer_ef = embedding_functions.SentenceTransformerEmbeddingFunction(
    model_name=EMBEDDING_MODEL_NAME)  # ИЗМЕНЕНО

# Создание клиента ChromaDB
client = chromadb.PersistentClient(path=CHROMA_DB_PATH)


def get_collection(label: str):
    """Возвращает или создает коллекцию для указанного label."""
    collection_name = f"label_{label}"  # Имя коллекции на основе label
    return client.get_or_create_collection(
        name=collection_name,
        embedding_function=sentence_transformer_ef,  # используем SentenceTransformer
    )


# --- Функция разбиения текста на абзацы ---
def split_text_into_paragraphs(text: str) -> List[str]:
    """
    Разбивает текст на абзацы.

    Args:
        text: Исходный текст.

    Returns:
        Список абзацев текста.
    """
    paragraphs = text.split("\n\n")  # Разделение по двойным переносам строк (типичный признак абзаца)
    return [p.strip() for p in paragraphs if p.strip()]  # Убираем пробелы и пустые абзацы


# --- Функция разбиения на предложения ---
def split_paragraph_into_sentences(paragraph: str) -> List[str]:
    """
    Разбивает абзац на предложения с использованием nltk.sent_tokenize.
    """
    try:
        nltk.data.find("tokenizers/punkt")
    except LookupError:
        nltk.download("punkt")
    return nltk.sent_tokenize(paragraph)


def split_text_semantically(text: str, chunk_size: int = 500, chunk_overlap: int = 50) -> List[str]:
    """
    Разбивает текст на семантические чанки с использованием Langchain.
    """
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", " ", ""],  # Порядок важен
        length_function=len,  # Используем стандартную функцию len
    )
    chunks = text_splitter.split_text(text)
    return chunks


# --- FastAPI App ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    Lifespan event handler.
    """
    # Загружаем данные из файлов при старте
    await load_telegram_messages_from_files()
    asyncio.create_task(telegram_message_collector())
    asyncio.create_task(periodic_save_task())  # Запускаем периодическое сохранение
    yield
    # Сохраняем данные перед завершением работы
    await save_all_telegram_messages_to_files()


app = FastAPI(lifespan=lifespan)


# --- Зависимость для аутентификации по токену ---
async def verify_token(authorization: Optional[str] = Header(None)):
    """
    Проверяет токен из заголовка Authorization (схема Bearer).
    """
    if not authorization:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Отсутствует заголовок Authorization",
            headers={"WWW-Authenticate": "Bearer"},
        )

    try:
        scheme, token = authorization.split()
        if scheme.lower() != "bearer":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Неверная схема аутентификации. Используйте Bearer.",
                headers={"WWW-Authenticate": "Bearer"},
            )
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверный формат заголовка Authorization",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if token != CHAT_TOKEN:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверный токен аутентификации",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return True


# --- Функции для извлечения текста из разных форматов файлов ---
def extract_text_from_pdf(file: io.BytesIO) -> str:
    """Извлекает текст из PDF-файла."""
    with pdfplumber.open(file) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""  # Добавляем or "" для обработки пустых страниц
    return text


def extract_text_from_txt(file: io.BytesIO) -> str:
    """Извлекает текст из TXT-файла."""
    return file.read().decode("utf-8")


def extract_text_from_docx(file: io.BytesIO) -> str:
    """Извлекает текст из DOCX-файла."""
    doc = docx.Document(file)
    text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
    return text


def extract_text_from_json(file: io.BytesIO) -> str:
    """Извлекает текст из JSON-файла (предполагается, что это просто текстовый JSON)."""
    try:
        data = json.load(file)
        return json.dumps(data, indent=4)  # Форматируем JSON для читаемости
    except json.JSONDecodeError:
        raise ValueError("Некорректный JSON формат.")


def extract_text_from_xlsx(file: io.BytesIO) -> str:
    """Извлекает текст из XLSX-файла."""
    workbook = openpyxl.load_workbook(file)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            row_values = [str(cell.value) for cell in row if
                          cell.value is not None]  # Преобразуем все значения в строки и обрабатываем None
            text += ", ".join(row_values) + "\n"
    return text


# --- API endpoints ---
@app.post("/add_document/", dependencies=[Depends(verify_token)])
async def add_document(
        file: UploadFile = File(...),
        label: str = "default",
        metadata: str = "{}",  # Передаем метаданные как строку JSON
        document_id: str = None  # Allow specifying the document_id in the query
):
    """Добавляет документ в векторное хранилище, извлекая текст из файла.

    Args:
        file: Файл для загрузки (поддерживаются форматы: pdf, txt, docx, json, xlsx).
        label: Label для документа.
        metadata: JSON-строка с метаданными.
        document_id:  Уникальный идентификатор документа (если не указан, будет сгенерирован).
    """
    try:
        file_content = await file.read()
        file_io = io.BytesIO(file_content)
        file_type = file.filename.split(".")[-1].lower()

        if file_type == "pdf":
            text = extract_text_from_pdf(file_io)
        elif file_type == "txt":
            text = extract_text_from_txt(file_io)
        elif file_type == "docx":
            text = extract_text_from_docx(file_io)
        elif file_type == "json":
            text = extract_text_from_json(file_io)
        elif file_type == "xlsx":
            text = extract_text_from_xlsx(file_io)
        else:
            raise HTTPException(status_code=400, detail="Неподдерживаемый формат файла")

        # Преобразуем метаданные из строки JSON в словарь
        try:
            metadata_dict = json.loads(metadata)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Некорректный формат метаданных JSON")

        if not document_id:
            document_id = file.filename  # Используем имя файла как ID по умолчанию, если не указан

        collection = get_collection(label)

        # Удаляем старые чанки документа, если они есть
        collection.delete(where={"source_document_id": document_id})

        chunks = split_text_semantically(text)
        chunk_ids = [f"{document_id}_{i}" for i in range(len(chunks))]

        metadatas = []
        for i in range(len(chunks)):
            chunk_metadata = metadata_dict.copy()
            chunk_metadata["source_document_id"] = document_id
            metadatas.append(chunk_metadata)

        collection.upsert(
            documents=chunks,
            metadatas=metadatas,
            ids=chunk_ids,
        )

        return {"message": f"Документ {file.filename} успешно обработан и добавлен с label {label} и ID {document_id}."}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/query", response_model=ContextResponse, dependencies=[Depends(verify_token)])
async def query_documents(query: Query):
    """Выполняет поиск документов в векторном хранилище."""
    all_results = []
    for label in query.labels:  # Итерируемся по списку labels
        try:
            collection = get_collection(label)  # Получаем коллекцию для label
            results = collection.query(
                query_texts=[query.text],
                n_results=query.n_results,
                where=query.where  # Фильтры по метаданным
            )
            """
                $eq – оператор "равно". Другие операторы:
                $ne: Не равно
                $gt: Больше
                $gte: Больше или равно
                $lt: Меньше
                $lte: Меньше или равно
                $in: Значение находится в списке значений (например, {"category": {"$in": ["blog", "news"]}})
                $nin: Значение отсутствует в списке значений
                $contains: Строка содержит подстроку (например, {"author": {"$contains": "Doe"}})
                $like: Строка соответствует шаблону SQL-стиля LIKE.
            """
            # Добавляем результаты в общий список
            for i in range(len(results['ids'][0])):
                all_results.append({
                    'id': results['ids'][0][i],
                    'document': results['documents'][0][i],
                    'metadata': results['metadatas'][0][i],
                    'distance': results['distances'][0][i] if 'distances' in results else None,  # Добавляем расстояние
                    'label': label  # Добавляем label
                })
        except Exception as e:
            logging.error(f"Ошибка при запросе label {label}: {e}")  # Логируем ошибку, но продолжаем
            # raise HTTPException(status_code=500, detail=str(e)) # Можно и выкидывать исключение, но лучше обработать все label

    # Возвращаем все результаты, объединенные в один список
    return ContextResponse(results=all_results)


@app.delete("/delete_document/{document_id}/{label}", dependencies=[Depends(verify_token)])
async def delete_document(document_id: str, label: str):
    """Удаляет документ или фрагмент документа из векторного хранилища по ID и label."""
    try:
        collection = get_collection(label)  # Получаем коллекцию для label
        # !ВАЖНО: Удаляем все фрагменты, связанные с исходным document_id
        collection.delete(where={"source_document_id": document_id})
        return {"message": f"Документ с ID {document_id} и все его фрагменты с label {label} успешно удалены"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    return {"status": "ok"}


# --- Telegram integration functions ---
import httpx


async def get_telegram_updates(offset: int = 0) -> List[Dict]:
    """
    Получает обновления от Telegram API.
    """
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
    params = {"offset": offset, "timeout": 60}
    ssl_context = ssl.create_default_context(cafile=certifi.where())

    try:
        async with httpx.AsyncClient(verify=ssl_context) as client:
            response = await client.get(url, params=params, timeout=60)
            response.raise_for_status()
            data = response.json()
            if data.get("ok"):
                return data["result"]
            else:
                logging.warning(f"Ошибка при получении обновлений: {data}")
                return []
    except httpx.HTTPStatusError as e:
        logging.warning(f"Ошибка HTTP: {e}")
        return []
    except httpx.RequestError as e:
        logging.error(f"Ошибка подключения к Telegram API: {e}")
        traceback.print_exc()
        return []
    except Exception as e:
        logging.error(f"Общая ошибка: {e}")
        traceback.print_exc()
        return []


async def get_chat_details(chat_id: int) -> Dict:
    """
    Получает детали чата из Telegram API.
    """
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getChat"
    params = {"chat_id": chat_id}
    ssl_context = ssl.create_default_context(cafile=certifi.where())

    try:
        async with httpx.AsyncClient(verify=ssl_context) as client:
            response = await client.get(url, params=params, timeout=60)
            response.raise_for_status()
            data = response.json()
            if data.get("ok"):
                return data["result"]
            else:
                logging.warning(f"Ошибка при получении информации о чате: {data}")
                return {}
    except httpx.HTTPStatusError as e:
        logging.warning(f"Ошибка HTTP при получении информации о чате: {e}")
        return {}
    except httpx.RequestError as e:
        logging.error(f"Ошибка подключения к Telegram API (getChat): {e}")
        traceback.print_exc()
        return {}
    except Exception as e:
        logging.error(f"Общая ошибка при получении информации о чате: {e}")
        traceback.print_exc()
        return {}


async def get_deal_from_hubspot(deal_id: str) -> Optional[Dict]:
    """Gets deal information from HubSpot API."""
    url = f"https://api.hubapi.com/crm/v3/objects/deals/{deal_id}?properties=dealname,company_ids,associations&associations=companies"
    headers = {"Authorization": f"Bearer {HUBSPOT_API_KEY}", "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            deal_data = response.json()
            return deal_data
    except httpx.HTTPStatusError as e:
        logging.warning(f"Ошибка HTTP при получении информации о сделке: {e}")
        return None
    except httpx.RequestError as e:
        logging.error(f"Ошибка подключения к HubSpot API (get deal): {e}")
        traceback.print_exc()
        return None
    except Exception as e:
        logging.error(f"Общая ошибка при получении информации о сделке: {e}")
        traceback.print_exc()
        return None


async def get_company_from_hubspot(company_id: str) -> Optional[Dict]:
    """Gets company information from HubSpot API."""
    url = f"https://api.hubapi.com/crm/v3/objects/companies/{company_id}?properties=name"
    headers = {"Authorization": f"Bearer {HUBSPOT_API_KEY}", "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            company_data = response.json()
            return company_data
    except httpx.HTTPStatusError as e:
        logging.warning(f"Ошибка HTTP при получении информации о компании: {e}")
        return None
    except httpx.RequestError as e:
        logging.error(f"Ошибка подключения к HubSpot API (get company): {e}")
        traceback.print_exc()
        return None
    except Exception as e:
        logging.error(f"Общая ошибка при получении информации о компании: {e}")
        traceback.print_exc()
        return None


daily_conversations: Dict[str, Dict[str, Dict[str, Dict]]] = {}


# Структура:
# {
#   "chat_title": {
#     "date": {
#       "conversation_id": [
#         {"message_id": 123, "text": "...", ...},
#         {"message_id": 456, "text": "...", ...},
#       ]
#     }
#   }
# }


async def save_telegram_messages():
    """Saves structured Telegram messages to ChromaDB, filtering by date."""
    global daily_conversations

    for chat_title, date_data in list(daily_conversations.items()):
        for date_str, conversations in list(date_data.items()):
            try:
                for conversation_id, conversation in list(conversations.items()):
                    label = f"telegram_sales"
                    base_document_id = conversation_id  # Use conversation_id as base document_id
                    document_id = base_document_id
                    combined_text = json.dumps(conversation, indent=2, ensure_ascii=False)  # Дампим весь словарь

                    try:
                        collection = get_collection(label)

                        # Check if document_id already exists
                        counter = 1
                        while True:
                            try:
                                # Проверяем наличие документа с текущим document_id
                                existing_results = collection.get(ids=[document_id], include=[])
                                if existing_results and existing_results['ids']:  # Если документ с таким ID существует
                                    counter += 1
                                    document_id = f"{base_document_id}_{counter}"  # Добавляем счетчик к ID
                                    continue  # Проверяем следующий ID
                                else:
                                    break  # ID свободен, выходим из цикла
                            except Exception as check_error:
                                logging.error(f"Ошибка при проверке ID {document_id}:{check_error}")
                                traceback.print_exc()
                                break  # Прерываем цикл, чтобы избежать бесконечного повторения

                        metadata = {  # todo: добавить в каждое сообщение имя партнера, добавить имя сделки
                            "source": "telegram",
                            "chat": chat_title,
                            "chat_id": conversation_id.split(":")[-2],
                            "origin_conversation_id": conversation_id,
                            "date": date_str,
                            "author_username": conversation.get("author", {}).get("username") if conversation else None,
                            "author_first_name": conversation.get("author", {}).get(
                                "first_name") if conversation else None,
                            "deal_id": conversation.get("deal_id"),
                            "deal_title": conversation.get("deal_title"),
                            "company_id": conversation.get("company_id"),
                            "company_title": conversation.get("company_title")
                        }
                        collection.upsert(
                            documents=[combined_text],
                            metadatas=[metadata],
                            ids=[document_id]
                        )
                        logging.info(
                            f"Conversation {conversation_id} from {chat_title} on {date_str} saved to ChromaDB with ID: {document_id}")

                        # Mark the file as committed
                        filename = generate_filename(chat_title, date_str, conversation_id)
                        committed_filename = generate_committed_filename(chat_title, date_str, conversation_id)
                        if os.path.exists(os.path.join(TEMP_STORAGE_PATH, filename)):
                            os.rename(os.path.join(TEMP_STORAGE_PATH, filename),
                                      os.path.join(TEMP_STORAGE_PATH, committed_filename))

                        del conversations[conversation_id]  # Remove from memory

                    except Exception as e:
                        logging.error(f"Error saving conversation {conversation_id} to ChromaDB: {e}")
                        traceback.print_exc()
                        continue  # Continue to the next conversation

                    if not conversations:
                        del date_data[date_str]
            except ValueError:
                logging.error(f"Неверный формат даты: {date_str}. Пропускаем.")
                continue

        if not date_data:
            del daily_conversations[chat_title]

    return "Сообщения сохранены в ChromaDB."


def generate_filename(chat_title: str, date: str, conversation_id: str) -> str:
    """Generates a filename for temporary storage."""
    return f"telegram_{chat_title}_{date}_{conversation_id}.json"


def generate_committed_filename(chat_title: str, date: str, conversation_id: str) -> str:
    """Generates a filename for committed storage."""
    return f"telegram_{chat_title}_{date}_{conversation_id}_commited.json"


async def save_all_telegram_messages_to_files():
    """Saves all telegram messages to temporary files."""
    global daily_conversations

    # Create the temporary storage directory if it doesn't exist
    if not os.path.exists(TEMP_STORAGE_PATH):
        os.makedirs(TEMP_STORAGE_PATH)

    for chat_title, date_data in daily_conversations.items():
        for date, conversations in date_data.items():
            for conversation_id, conversation in conversations.items():
                filename = generate_filename(chat_title, date, conversation_id)
                filepath = os.path.join(TEMP_STORAGE_PATH, filename)
                try:
                    async with aiofiles.open(filepath, 'w', encoding='utf-8') as f:
                        await f.write(
                            json.dumps(conversation, indent=2, ensure_ascii=False))  # Сохраняем словарь conversation
                    logging.info(
                        f"Conversation {conversation_id} from {chat_title} on {date} saved to file: {filename}")
                except Exception as e:
                    logging.error(f"Error saving conversation {conversation_id} to file {filename}: {e}")
                    traceback.print_exc()


async def sync_save_all_telegram_messages_to_files():
    """Saves all telegram messages to temporary files."""
    global daily_conversations

    # Create the temporary storage directory if it doesn't exist
    if not os.path.exists(TEMP_STORAGE_PATH):
        os.makedirs(TEMP_STORAGE_PATH)

    for chat_title, date_data in daily_conversations.items():
        for date, conversations in date_data.items():
            for conversation_id, conversation in conversations.items():
                filename = generate_filename(chat_title, date, conversation_id)
                filepath = os.path.join(TEMP_STORAGE_PATH, filename)
                try:
                    async with aiofiles.open(filepath, 'w', encoding='utf-8') as f:
                        await f.write(json.dumps(conversation, indent=2, ensure_ascii=False))
                    logging.info(
                        f"Conversation {conversation_id} from {chat_title} on {date} saved to file: {filename}")
                except Exception as e:
                    logging.error(f"Error saving conversation {conversation_id} to file {filename}: {e}")
                    traceback.print_exc()


async def load_telegram_messages_from_files():
    """Loads telegram messages from temporary files."""
    global daily_conversations

    # Create the temporary storage directory if it doesn't exist
    if not os.path.exists(TEMP_STORAGE_PATH):
        os.makedirs(TEMP_STORAGE_PATH)

    # Iterate through files in the temporary storage directory
    for filename in os.listdir(TEMP_STORAGE_PATH):
        if filename.endswith(".json") and not filename.endswith("_commited.json"):
            filepath = os.path.join(TEMP_STORAGE_PATH, filename)
            try:
                async with aiofiles.open(filepath, 'r', encoding='utf-8') as f:
                    conversation = json.loads(await f.read())

                # Extract chat_title, date, and conversation_id from the filename
                parts = filename.split("_")
                chat_title = parts[1]
                date = parts[2]
                conversation_id = parts[3].split(".")[0]

                # Load the conversation into the daily_conversations dictionary
                if chat_title not in daily_conversations:
                    daily_conversations[chat_title] = {}
                if date not in daily_conversations[chat_title]:
                    daily_conversations[chat_title][date] = {}
                daily_conversations[chat_title][date][conversation_id] = conversation  # Store the conversation directly

                logging.info(f"Conversation {conversation_id} from file {filename} loaded into memory.")

            except Exception as e:
                logging.error(f"Error loading conversation from file {filename}: {e}")
                traceback.print_exc()


async def load_offset():
    """Загружает offset из файла асинхронно."""
    try:
        async with aiofiles.open(OFFSET_FILE, "r") as f:
            content = await f.read()
            return int(content)
    except FileNotFoundError:
        return 0  # Начинаем с 0, если файл не существует
    except ValueError:
        print("Ошибка: Некорректный offset в файле. Начинаем с 0.")
        return 0


async def save_offset(offset: int):
    """Сохраняет offset в файл асинхронно."""
    try:
        async with aiofiles.open(OFFSET_FILE, "w") as f:
            await f.write(str(offset))
    except Exception as e:
        print(f"Ошибка при сохранении offset в файл: {e}")


deal_cache = {}  # Initialize the deal cache


async def telegram_message_collector():
    """Collects Telegram messages and stores them in structured format."""
    global daily_conversations
    offset = await load_offset()  # Загружаем offset при старте и в каждой итерации, на всякий случай
    logging.info(f"Загружен offset: {offset}")

    while True:
        updates = await get_telegram_updates(offset=offset)
        if updates:
            for update in updates:
                offset = update["update_id"] + 1
                if "message" in update:
                    message = update["message"]
                    chat_id = message["chat"]["id"]
                    # Get chat details
                    chat_details = await get_chat_details(chat_id)
                    chat_title = chat_details.get("title", f"ChatID_{chat_id}") if chat_details else f"ChatID_{chat_id}"
                    chat_description = chat_details.get("description", "") if chat_details else ""

                    # Extract deal ID from chat title or description
                    deal_id = None
                    match = re.search(r"(\d{11})", chat_title + chat_description)
                    if match:
                        deal_id = match.group(1)

                    date = datetime.fromtimestamp(message["date"]).strftime("%Y-%m-%d")

                    author_username = message["from"].get("username")
                    author_first_name = message["from"].get("first_name")

                    text = message.get("text", message.get("caption", None))
                    message_id = message.get("message_id")  # Extract the message id
                    reply_to_message = message.get("reply_to_message")  # Extract reply_to_message
                    conversation_message_start_id = message.get('message_thread_id',
                                                                message_id)  # message_thread_id for forums, message_id otherwise
                    logging.info(f"Received message from {chat_title} on {date}: {text}")
                    human_readable_date = datetime.fromtimestamp(message["date"]).strftime('%Y-%m-%d %H:%M:%S')

                    deal_title = None
                    company_id = None
                    company_title = None

                    # Enrich message with deal and company information
                    if deal_id:
                        if deal_id in deal_cache:
                            deal_data = deal_cache[deal_id]  # Get from cache
                            deal_title = deal_data.get("deal_title")
                            company_id = deal_data.get("company_id")
                            company_title = deal_data.get("company_title")
                        else:
                            deal_data = await get_deal_from_hubspot(deal_id)
                            if deal_data:
                                deal_title = deal_data["properties"].get("dealname")
                                company_ids = list(
                                    set([i['id'] for i in deal_data['associations']['companies']['results']]))
                                if company_ids and len(company_ids) > 0:
                                    company_id = company_ids[0]  # Assuming one company for deal
                                    company_data = await get_company_from_hubspot(company_id)
                                    if company_data:
                                        company_title = company_data["properties"].get("name")

                                # Store in cache
                                deal_cache[deal_id] = {
                                    "deal_title": deal_title,
                                    "company_id": company_id,
                                    "company_title": company_title
                                }

                    if text:
                        message_data = {
                            "message_id": message_id,
                            "date": human_readable_date,
                            "text": text,
                            "conversation_message_start_id": conversation_message_start_id,
                            "author": {
                                "username": author_username,
                                "first_name": author_first_name
                            },
                            "deal_id": deal_id,
                            "deal_title": deal_title,
                            "company_id": company_id,
                            "company_title": company_title,
                        }

                        conversation_id = f"{chat_title}{chat_id}:{conversation_message_start_id}"

                        if chat_title not in daily_conversations:
                            daily_conversations[chat_title] = {}
                        if date not in daily_conversations[chat_title]:
                            daily_conversations[chat_title][date] = {}

                        if conversation_id not in daily_conversations[chat_title][date]:
                            daily_conversations[chat_title][date][
                                conversation_id] = message_data  # Сохраняем сообщение напрямую
                        else:
                            # Handle replies to messages
                            if reply_to_message:
                                # Find the parent message and nest the reply.
                                found_parent = False
                                conversation = daily_conversations[chat_title][date][conversation_id]
                                if find_and_nest_reply([conversation], reply_to_message,
                                                       message_data):  # Обратите внимание: передаем conversation в списке
                                    found_parent = True

                                if not found_parent:
                                    # Add as new if not found (rare case, but must be handled)
                                    logging.warning(
                                        f"Parent message not found for reply message_id: {message_id}, reply_to_message_id: {reply_to_message.get('message_id')}")

            await save_offset(offset)  # Сохраняем offset после обработки пакета
            logging.info(f"Сохранен offset: {offset}")
        else:
            await asyncio.sleep(1)  # Избегаем частых запросов, если нет обновлений

        # Periodic save
        now = datetime.now(timezone.utc) + timedelta(hours=3)
        if now.hour == 0 and now.minute == 0:
            await save_telegram_messages()
        await asyncio.sleep(5)


def find_and_nest_reply(conversation: List[Dict], reply_to_message: Dict,
                        message_data: Dict) -> bool:  # Оставляем List[Dict] для совместимости со старым кодом, т.к. рекурсивно вызываем сами себя
    """
    Recursively searches for the parent message and nests the reply.
    Returns True if parent found and nested, False otherwise.
    """
    msg = conversation[0]  # Берем единственный элемент из списка
    if msg.get("message_id") == reply_to_message.get("message_id"):
        # Found the parent message, nest the reply
        if "replies" not in msg:
            msg["replies"] = []
        msg["replies"].append(message_data)  # Nest message_data
        return True
    if "replies" in msg:
        # Оборачиваем replies в список, чтобы рекурсивно вызывать функцию
        if find_and_nest_reply(msg["replies"], reply_to_message, message_data):
            return True  # Parent found in nested replies
    return False  # Parent not found in this conversation


@app.post("/force_save/", response_model=ForceSaveResponse, dependencies=[Depends(verify_token)])
async def force_save_messages():
    """
    Endpoint to force save the telegram messages to ChromaDB.
    """
    await save_all_telegram_messages_to_files()
    message = await save_telegram_messages()
    return ForceSaveResponse(message=message)


# --- Signal Handling and atexit ---
async def handle_exit():
    logging.info("Завершение работы: сохранение данных...")
    await sync_save_all_telegram_messages_to_files()
    logging.info("Данные сохранены.")


# Register the exit handler
# atexit.register(handle_exit)

def register_signal_handlers():
    """Registers signal handlers for SIGINT and SIGTERM."""

    async def signal_handler(sig, frame):
        logging.info(f"Received signal {sig}. Exiting...")
        try:
            await save_all_telegram_messages_to_files()  # Save data before exiting
            await save_telegram_messages()
            logging.info("Data saved successfully.")
        except Exception as e:
            logging.error(f"Error saving data during shutdown: {e}")
            traceback.print_exc()
        finally:
            sys.exit(0)

    def handle_signal(sig, frame):
        loop = asyncio.get_event_loop()
        loop.run_until_complete(signal_handler(sig, frame))

    signal.signal(signal.SIGINT, handle_signal)  # Ctrl+C
    signal.signal(signal.SIGTERM, handle_signal)  # Termination signal (e.g., from Docker)


async def periodic_save_task():
    """Периодически сохраняет данные."""
    while True:
        logging.info("Периодическое сохранение данных...")
        try:
            await sync_save_all_telegram_messages_to_files()
            await save_telegram_messages()  # Сохраняем только старые сообщения
            logging.info("Периодическое сохранение успешно.")
        except Exception as e:
            logging.error(f"Ошибка при периодическом сохранении: {e}")
            traceback.print_exc()
        await asyncio.sleep(SAVE_INTERVAL_SECONDS)


# --- Main ---
if __name__ == "__main__":
    register_signal_handlers()

    uvicorn.run(app, host="0.0.0.0", port=8001)
